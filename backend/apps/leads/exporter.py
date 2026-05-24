import csv
import os
import uuid
from datetime import timedelta

from django.conf import settings
from django.db.models import Q
from django.http import HttpResponse, FileResponse
from django.utils import timezone

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from apps.searches.models import Search
from apps.usage.services import get_or_create_usage

from .models import Lead, LeadList, ExportHistory


EXPORT_HEADERS = [
    "Business Name",
    "Keyword",
    "Location",
    "Category",
    "Phone",
    "Email 1",
    "Email 2",
    "Email 3",
    "Email Confidence",
    "Website",
    "Domain",
    "Website Status",
    "HTTP Status",
    "Website Error",
    "Website Platform",
    "Is Broken Website",
    "Is Social Only",
    "Is Free Builder",
    "Facebook",
    "Instagram",
    "LinkedIn",
    "YouTube",
    "TikTok",
    "Address",
    "City",
    "State",
    "Pincode",
    "Country",
    "Rating",
    "Review Count",
    "Map Link",
    "Status",
    "Lead Score",
    "Opportunity Score",
    "Opportunity Reason",
    "Tags",
    "Enrichment Status",
    "Created At",
]


AI_EXPORT_HEADERS = [
    "AI Priority",
    "AI Summary",
    "AI Target Offer",
    "AI Campaign Goal",
    "AI Tone",
    "AI Target Audience",
    "AI Outreach Channel",
    "AI Suggested Offer",
    "AI Offer Reason",
    "AI Best Channel",
    "AI Channel Reason",
    "AI First Line",
    "AI Email Subject",
    "AI Email Body",
    "AI Follow Up 1",
    "AI Follow Up 2",
    "AI Follow Up 3",
    "AI Facebook Message",
    "AI LinkedIn Message",
    "AI WhatsApp Message",
    "AI Website Weakness",
    "AI Local SEO Opportunity",
    "AI Score Explanation",
    "AI Provider",
    "AI Model",
    "AI Generated At",
]


def normalize_boolean(value):
    if value is None:
        return None

    if isinstance(value, bool):
        return value

    if isinstance(value, str):
        value = value.strip().lower()

        if value in ["true", "1", "yes", "on"]:
            return True

        if value in ["false", "0", "no", "off"]:
            return False

    return None


def normalize_int_list(value):
    if not value:
        return []

    if isinstance(value, list):
        items = value
    elif isinstance(value, str):
        items = value.split(",")
    else:
        items = [value]

    final_ids = []

    for item in items:
        try:
            final_ids.append(int(item))
        except Exception:
            continue

    return list(set(final_ids))


def build_export_headers(include_ai=False):
    include_ai = normalize_boolean(include_ai) is True

    if include_ai:
        return EXPORT_HEADERS + AI_EXPORT_HEADERS

    return EXPORT_HEADERS


def get_export_base_queryset(user):
    return (
        Lead.objects.filter(search__user=user)
        .select_related("search")
        .select_related("search", "ai_insight")  # .prefetch_related("ai_insight")
    )


def apply_export_filters(
    leads,
    search_id=None,
    keyword=None,
    location=None,
    status=None,
    rating=None,
    has_email=None,
    has_website=None,
    website_status=None,
    enrichment_status=None,
    min_score=None,
):
    if search_id:
        leads = leads.filter(search_id=search_id)

    if keyword:
        leads = leads.filter(keyword__icontains=keyword)

    if location:
        leads = leads.filter(location__icontains=location)

    if status:
        leads = leads.filter(status=status)

    if website_status:
        leads = leads.filter(website_status=website_status)

    if enrichment_status:
        leads = leads.filter(enrichment_status=enrichment_status)

    if rating:
        try:
            rating_value = float(rating)
            leads = leads.filter(rating__gte=rating_value)
        except Exception:
            pass

    if min_score:
        try:
            min_score_value = int(min_score)
            leads = leads.filter(opportunity_score__gte=min_score_value)
        except Exception:
            pass

    has_email_value = normalize_boolean(has_email)
    has_website_value = normalize_boolean(has_website)

    if has_email_value is True:
        leads = leads.exclude(email_1__isnull=True).exclude(email_1="")

    if has_email_value is False:
        leads = leads.filter(Q(email_1__isnull=True) | Q(email_1=""))

    if has_website_value is True:
        leads = leads.filter(has_website=True)

    if has_website_value is False:
        leads = leads.filter(has_website=False)

    return leads


def build_export_queryset(
    user,
    search_id=None,
    lead_list_id=None,
    selected_lead_ids=None,
    keyword=None,
    location=None,
    status=None,
    rating=None,
    has_email=None,
    has_website=None,
    website_status=None,
    enrichment_status=None,
    min_score=None,
):
    leads = get_export_base_queryset(user)

    selected_lead_ids = normalize_int_list(selected_lead_ids)

    if selected_lead_ids:
        leads = leads.filter(id__in=selected_lead_ids)

    if lead_list_id:
        lead_list = LeadList.objects.filter(id=lead_list_id, user=user).first()

        if not lead_list:
            return Lead.objects.none()

        leads = leads.filter(list_items__lead_list=lead_list)

    leads = apply_export_filters(
        leads=leads,
        search_id=search_id,
        keyword=keyword,
        location=location,
        status=status,
        rating=rating,
        has_email=has_email,
        has_website=has_website,
        website_status=website_status,
        enrichment_status=enrichment_status,
        min_score=min_score,
    )

    return leads.distinct().order_by("-created_at")


def get_lead_ai_insight(lead):
    try:
        return lead.ai_insight
    except Exception:
        return None


def lead_to_ai_row(lead):
    insight = get_lead_ai_insight(lead)

    if not insight:
        return [""] * len(AI_EXPORT_HEADERS)

    return [
        insight.ai_priority,
        insight.ai_summary,
        insight.target_offer,
        insight.campaign_goal,
        insight.tone,
        insight.target_audience,
        insight.outreach_channel,
        insight.ai_suggested_offer,
        insight.ai_offer_reason,
        insight.ai_best_channel,
        insight.ai_channel_reason,
        insight.ai_first_line,
        insight.ai_email_subject,
        insight.ai_email_body,
        insight.ai_followup_1,
        insight.ai_followup_2,
        getattr(insight, "ai_followup_3", ""),
        insight.ai_facebook_message,
        getattr(insight, "ai_linkedin_message", ""),
        insight.ai_whatsapp_message,
        insight.ai_website_weakness,
        insight.ai_local_seo_opportunity,
        insight.ai_score_explanation,
        insight.provider,
        insight.model_name,
        insight.generated_at,
    ]


def lead_to_row(lead, include_ai=False):
    row = [
        lead.name,
        lead.keyword,
        lead.location,
        lead.category,
        lead.phone,
        lead.email_1,
        lead.email_2,
        lead.email_3,
        lead.email_confidence,
        lead.website,
        lead.domain,
        lead.website_status,
        lead.website_http_status,
        lead.website_error,
        lead.website_platform,
        lead.is_broken_website,
        lead.is_social_only,
        lead.is_free_builder,
        lead.facebook_url,
        lead.instagram_url,
        lead.linkedin_url,
        lead.youtube_url,
        lead.tiktok_url,
        lead.address,
        lead.city,
        lead.state,
        lead.pincode,
        lead.country,
        lead.rating,
        lead.review_count,
        lead.map_link,
        lead.status,
        lead.lead_score,
        lead.opportunity_score,
        lead.opportunity_reason,
        ", ".join(lead.tags or []),
        lead.enrichment_status,
        lead.created_at,
    ]

    include_ai = normalize_boolean(include_ai) is True

    if include_ai:
        row.extend(lead_to_ai_row(lead))

    return row


def get_export_scope(search_id=None, lead_list_id=None, selected_lead_ids=None):
    selected_lead_ids = normalize_int_list(selected_lead_ids)

    if selected_lead_ids:
        return ExportHistory.SCOPE_SELECTED

    if lead_list_id:
        return ExportHistory.SCOPE_LEAD_LIST

    if search_id:
        return ExportHistory.SCOPE_SEARCH

    return ExportHistory.SCOPE_ALL


def get_export_file_name(export_history):
    extension = export_history.export_type

    if export_history.export_scope == ExportHistory.SCOPE_SEARCH and export_history.search_id:
        base_name = f"search_{export_history.search_id}_leads"

    elif export_history.export_scope == ExportHistory.SCOPE_LEAD_LIST and export_history.lead_list_id:
        base_name = f"lead_list_{export_history.lead_list_id}"

    elif export_history.export_scope == ExportHistory.SCOPE_SELECTED:
        base_name = "selected_leads"

    else:
        base_name = "leads"

    unique_suffix = uuid.uuid4().hex[:8]

    return f"{base_name}_{unique_suffix}.{extension}"


def get_export_file_path(user, file_name):
    base_dir = settings.EXPORTS_ROOT / f"user_{user.id}"
    os.makedirs(base_dir, exist_ok=True)

    return base_dir / file_name


def create_export_history(
    user,
    export_type=ExportHistory.EXPORT_TYPE_CSV,
    search_id=None,
    lead_list_id=None,
    selected_lead_ids=None,
    filters=None,
    total_rows=0,
    status=ExportHistory.STATUS_PENDING,
):
    usage = get_or_create_usage(user)

    expires_at = timezone.now() + timedelta(days=usage.export_retention_days)

    search = None
    lead_list = None

    if search_id:
        search = Search.objects.filter(id=search_id, user=user).first()

    if lead_list_id:
        lead_list = LeadList.objects.filter(id=lead_list_id, user=user).first()

    export_scope = get_export_scope(
        search_id=search_id,
        lead_list_id=lead_list_id,
        selected_lead_ids=selected_lead_ids,
    )

    selected_lead_ids = normalize_int_list(selected_lead_ids)

    return ExportHistory.objects.create(
        user=user,
        search=search,
        lead_list=lead_list,
        export_type=export_type,
        export_scope=export_scope,
        status=status,
        filters=filters or {},
        selected_lead_ids=selected_lead_ids,
        total_rows=total_rows,
        expires_at=expires_at,
    )


def write_csv_file(leads, file_path, include_ai=False):
    total = 0
    include_ai = normalize_boolean(include_ai) is True

    with open(file_path, "w", newline="", encoding="utf-8-sig") as csv_file:
        writer = csv.writer(csv_file)
        writer.writerow(build_export_headers(include_ai=include_ai))

        for lead in leads.iterator(chunk_size=500):
            writer.writerow(lead_to_row(lead, include_ai=include_ai))
            total += 1

    return total


def write_xlsx_file(leads, file_path, include_ai=False):
    include_ai = normalize_boolean(include_ai) is True

    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)

    ws.append(build_export_headers(include_ai=include_ai))

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    total = 0

    for lead in leads.iterator(chunk_size=500):
        ws.append(lead_to_row(lead, include_ai=include_ai))
        total += 1

    for column_cells in ws.columns:
        max_length = 12
        column_letter = column_cells[0].column_letter

        for cell in column_cells:
            try:
                value_length = len(str(cell.value or ""))

                if value_length > max_length:
                    max_length = min(value_length, 45)

            except Exception:
                pass

        ws.column_dimensions[column_letter].width = max_length + 2

    ws.freeze_panes = "A2"
    wb.save(file_path)

    return total


def generate_export_file(export_history):
    export_history.status = ExportHistory.STATUS_RUNNING
    export_history.started_at = timezone.now()
    export_history.error_message = None
    export_history.save(
        update_fields=[
            "status",
            "started_at",
            "error_message",
            "updated_at",
        ]
    )

    filters = export_history.filters or {}
    include_ai = normalize_boolean(filters.get("include_ai")) is True

    leads = build_export_queryset(
        user=export_history.user,
        search_id=filters.get("search_id"),
        lead_list_id=filters.get("lead_list_id"),
        selected_lead_ids=export_history.selected_lead_ids,
        keyword=filters.get("keyword"),
        location=filters.get("location"),
        status=filters.get("status"),
        rating=filters.get("rating"),
        has_email=filters.get("has_email"),
        has_website=filters.get("has_website"),
        website_status=filters.get("website_status"),
        enrichment_status=filters.get("enrichment_status"),
        min_score=filters.get("min_score"),
    )

    file_name = get_export_file_name(export_history)
    file_path = get_export_file_path(export_history.user, file_name)

    if export_history.export_type == ExportHistory.EXPORT_TYPE_XLSX:
        total_rows = write_xlsx_file(
            leads,
            file_path,
            include_ai=include_ai,
        )
    else:
        total_rows = write_csv_file(
            leads,
            file_path,
            include_ai=include_ai,
        )

    file_size = os.path.getsize(file_path) if os.path.exists(file_path) else 0

    export_history.status = ExportHistory.STATUS_COMPLETED
    export_history.total_rows = total_rows
    export_history.file_name = file_name
    export_history.file_path = str(file_path)
    export_history.file_size_bytes = file_size
    export_history.completed_at = timezone.now()
    export_history.save(
        update_fields=[
            "status",
            "total_rows",
            "file_name",
            "file_path",
            "file_size_bytes",
            "completed_at",
            "updated_at",
        ]
    )

    return export_history


def mark_export_failed(export_history, error_message):
    export_history.status = ExportHistory.STATUS_FAILED
    export_history.error_message = str(error_message)[:1000]
    export_history.completed_at = timezone.now()
    export_history.save(
        update_fields=[
            "status",
            "error_message",
            "completed_at",
            "updated_at",
        ]
    )

    return export_history


def export_leads_csv(
    user,
    search_id=None,
    keyword=None,
    location=None,
    status=None,
    rating=None,
    has_email=None,
    has_website=None,
    include_ai=False,
):
    """
    Backward-compatible immediate CSV export endpoint.

    Optional:
    include_ai=true adds AI insight/personalization columns.
    """

    include_ai = normalize_boolean(include_ai) is True

    leads = build_export_queryset(
        user=user,
        search_id=search_id,
        keyword=keyword,
        location=location,
        status=status,
        rating=rating,
        has_email=has_email,
        has_website=has_website,
    )

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="leads.csv"'

    writer = csv.writer(response)
    writer.writerow(build_export_headers(include_ai=include_ai))

    for lead in leads.iterator(chunk_size=500):
        writer.writerow(lead_to_row(lead, include_ai=include_ai))

    return response


def get_export_download_response(export_history):
    if not export_history.file_path:
        return None

    if not os.path.exists(export_history.file_path):
        return None

    content_type = "text/csv"

    if export_history.export_type == ExportHistory.EXPORT_TYPE_XLSX:
        content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    return FileResponse(
        open(export_history.file_path, "rb"),
        as_attachment=True,
        filename=export_history.file_name,
        content_type=content_type,
    )